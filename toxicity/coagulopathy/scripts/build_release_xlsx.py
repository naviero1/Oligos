#!/usr/bin/env python3
"""Build the Phase 2 dataset release workbook from the canonical CSVs.

    python3 toxicity/coagulopathy/scripts/build_release_xlsx.py

The Challenge asks for the dataset as "a data dictionary and schema documenting all
metadata, and access to the raw data ... by including a data file in Excel (or similar
format)". This produces that single file. The CSVs in data/ remain canonical; this
workbook is generated from them and is never hand-edited.

Summary figures are written as VALUES, not formulas. The sibling CNS release used live
COUNTA/COUNTIF formulas, which read back as empty cells in any tool that does not
recalculate on open (pandas.read_excel among them). Values avoid that trap; the numbers are
recomputed here on every build, so they cannot drift from the data.
"""
import csv, os, sys
from collections import Counter

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl")

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = os.path.join(ROOT, "data")
OUT = os.path.join(ROOT, "OligoTox-Coagulopathy_Dataset.xlsx")

HDR = Font(bold=True, color="FFFFFF")
FILL = PatternFill("solid", fgColor="1F3864")
TITLE = Font(bold=True, size=13)


def load(n):
    with open(os.path.join(DATA, n), newline="", encoding="utf-8") as fh:
        return list(csv.DictReader(fh))


def sheet(wb, name, rows, freeze="A2"):
    ws = wb.create_sheet(name)
    if not rows:
        return ws
    cols = list(rows[0].keys())
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        ws.cell(1, c).font = HDR
        ws.cell(1, c).fill = FILL
    for r in rows:
        ws.append([r.get(c, "") for c in cols])
    for i, c in enumerate(cols, 1):
        w = max(len(c), *(len(str(r.get(c, ""))[:60]) for r in rows[:400])) + 2
        ws.column_dimensions[get_column_letter(i)].width = min(w, 60)
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions
    return ws


def _runs(pairs):
    """Compact a per-position attribute into runs: [(1,'2\'-MOE'),...] -> "1-5:2'-MOE; 6-15:DNA"."""
    if not pairs:
        return "NOT_REPORTED"
    out, start, prev_pos, prev_val = [], pairs[0][0], pairs[0][0], pairs[0][1]
    for pos, val in pairs[1:]:
        if val != prev_val or pos != prev_pos + 1:
            out.append(f"{start}-{prev_pos}:{prev_val}" if start != prev_pos else f"{start}:{prev_val}")
            start, prev_val = pos, val
        prev_pos = pos
    out.append(f"{start}-{prev_pos}:{prev_val}" if start != prev_pos else f"{start}:{prev_val}")
    return "; ".join(out)


def sequence_status(o):
    """Why a compound has no sequence -- so a gap is actionable rather than blank.

    'published' means this release carries it. The other four say what kind of gap it is,
    which is the difference between a compound whose sequence CANNOT exist as one string
    and one that simply has not been recovered yet."""
    if o["sequence_base"] not in ("NOT_REPORTED", "NOT_APPLICABLE", ""):
        return "published"
    if o["oligo_class"] == "polydisperse_ssDNA":
        return "not_applicable_polydisperse_mixture"
    if o["modality"] == "double_stranded_siRNA" or "duplex" in str(o["sequence_note"]).lower() \
       or "sense" in str(o["sequence_note"]).lower():
        return "not_applicable_duplex_two_strands"
    if "pprov" in str(o["max_phase"]) or "arket" in str(o["max_phase"]):
        return "recoverable_from_WHO_INN_nomenclature"
    return "not_published_by_source"

def modification_maps(M):
    """Per-oligo compact chemistry maps, built from the per-position table."""
    by = {}
    for r in M:
        by.setdefault(r["oligo_id"], []).append(r)
    out = {}
    for oid, rows in by.items():
        rows = sorted(rows, key=lambda r: int(r["position"]) if str(r["position"]).isdigit() else 0)
        out[oid] = {
            "sugar_map": _runs([(int(r["position"]), r["sugar_mod"]) for r in rows if str(r["position"]).isdigit()]),
            "backbone_map": _runs([(int(r["position"]), r["backbone_linkage_3p"]) for r in rows if str(r["position"]).isdigit()]),
            "n_positions": len(rows),
            "per_position": " ".join(f"{r['position']}{r['nucleobase']}({r['sugar_mod']}/{r['backbone_linkage_3p']})" for r in rows),
        }
    return out


def toxicity_rollup(D):
    """Per-oligo toxicity summary. worst_grade is the maximum ASSIGNED grade; an oligo whose
    rows are all ungraded gets NOT_REPORTED, never 0 -- absence of a grade is not safety."""
    by = {}
    for r in D:
        by.setdefault(r["oligo_id"], []).append(r)
    out = {}
    for oid, rows in by.items():
        gr = [int(r["coag_tox_grade"]) for r in rows if r["coag_tox_grade"] in ("0", "1", "2", "3")]
        sev = [r["severity_stated_by_source"] for r in rows
               if r["severity_stated_by_source"] not in ("NOT_REPORTED", "NOT_APPLICABLE", "")]
        sg = [r["source_stated_grade"] for r in rows if r["source_stated_grade"] != "NOT_APPLICABLE"]
        out[oid] = {
            "worst_coag_tox_grade": str(max(gr)) if gr else "NOT_REPORTED",
            "grade_distribution_0_1_2_3": "/".join(str(sum(1 for g in gr if g == k)) for k in range(4)) if gr else "NOT_REPORTED",
            "n_graded_rows": len(gr),
            "n_rows_flagged_grade_caveat": sum(1 for r in rows if r["grade_caveat"] == "within_reference_range_resolution"),
            "max_source_stated_grade": max(sg) if sg else "NOT_APPLICABLE",
            "severity_in_source_words": (sev[0][:300] if sev else "NOT_REPORTED"),
            "n_human_rows": sum(1 for r in rows if r["species_class"] == "human"),
            "n_animal_rows": sum(1 for r in rows if r["species_class"] == "animal"),
            "on_target_rows": sum(1 for r in rows if r["on_target_effect"] == "TRUE"),
            "unintended_toxicity_rows": sum(1 for r in rows if r["unintended_toxicity"] == "TRUE"),
        }
    return out


def main():
    S, O, M, D = load("sources.csv"), load("oligos.csv"), load("modifications.csv"), load("measurements.csv")
    wb = Workbook()
    wb.remove(wb.active)

    # ---- README ----------------------------------------------------------
    ws = wb.create_sheet("README")
    NR = "NOT_REPORTED"
    lines = [
        ("OligoTox-Coagulopathy", TITLE),
        ("Coagulation toxicity of oligonucleotide therapeutics. NIH/NCATS Oligonucleotide Toxicity Open Data Challenge, Phase 2.", None),
        ("", None),
        ("Licence: Creative Commons Attribution 4.0 (CC BY 4.0) for the curated tables.", None),
        ("Underlying third-party full texts are referenced, not redistributed; per-row terms are in the 'redistribution' column.", None),
        ("", None),
        ("SHEETS", TITLE),
        ("Summary            headline counts, computed from the data sheets at build time", None),
        ("data_dictionary    every column in every table, with its definition", None),
        ("sources            the provenance registry - one row per source document", None),
        ("oligos             one row per oligonucleotide - the predictor variables", None),
        ("measurements       one row per measured outcome - the response variables", None),
        ("modifications      one row per NUCLEOTIDE POSITION - the per-position chemistry", None),
        ("", None),
        ("HOW THE TABLES JOIN", TITLE),
        ("measurements.oligo_id  -> oligos.oligo_id      (many measurements per oligonucleotide)", None),
        ("measurements.source_id -> sources.source_id", None),
        ("modifications.oligo_id -> oligos.oligo_id      (one row per position, 5' to 3')", None),
        ("", None),
        ("MISSING VALUES - READ THIS BEFORE ANALYSING", TITLE),
        ("NOT_REPORTED    the source does not report this value. It has NOT been estimated, imputed or filled from background knowledge.", None),
        ("NOT_APPLICABLE  the field has no meaning for this row (a dose for an in-vitro spike-in; a 5'->3' string for a duplex or a polydisperse mixture).", None),
        ("Cells are never blank and never zero-as-missing.", None),
        ("", None),
        ("TWO AXES, NOT ONE - THE MOST IMPORTANT THING ON THIS PAGE", TITLE),
        ("Most rows here are ON-TARGET anticoagulant pharmacology, not toxicity: the compounds with published clotting", None),
        ("numbers are largely the ones designed to change clotting. on_target_effect and unintended_toxicity are separate", None),
        ("flags and BOTH may be true on one row. A model trained across them without the flags learns that anticoagulants", None),
        ("prolong aPTT - true, circular, and useless for safety prediction.", None),
        ("", None),
        ("HUMAN VERSUS ANIMAL", TITLE),
        ("species_class (human / animal / not_determined) and human_system carry this distinction, NOT study_type.", None),
        ("A purified-protein assay counts as a human system when the proteins are human; species_class_basis records how", None),
        ("each row was classified.", None),
        ("", None),
        ("GRADES ARE PROVISIONAL", TITLE),
        ("coag_tox_grade is assigned mechanically from a control-referenced ratio by CTCAE v5.0 cut-offs, and only for the", None),
        ("readouts CTCAE defines. CTCAE grades against the upper limit of normal; these sources publish a control mean, so", None),
        ("grades resting on a ratio of 1.0-1.2x carry grade_caveat = within_reference_range_resolution. FILTER ON THAT", None),
        ("BEFORE TREATING GRADE 1 AS A FINDING. No grade has had subject-matter review.", None),
    ]
    for text, font in lines:
        ws.append([text])
        if font:
            ws.cell(ws.max_row, 1).font = font
    ws.column_dimensions["A"].width = 130

    # ---- Summary ---------------------------------------------------------
    ws = wb.create_sheet("Summary")
    def block(title, pairs):
        ws.append([title]); ws.cell(ws.max_row, 1).font = TITLE
        for k, v in pairs:
            ws.append([k, v])
        ws.append([])
    seq = [r for r in O if r["sequence_base"] not in (NR, "NOT_APPLICABLE", "")]
    gr = Counter(r["coag_tox_grade"] for r in D)
    sc = Counter(r["species_class"] for r in D)
    st = Counter(r["study_type"] for r in D)
    rc = Counter(r["readout_category"] for r in D)
    block("OligoTox-Coagulopathy - release summary", [("Every figure below is computed from the data sheets at build time.", "")])
    block("Counts", [("Oligonucleotides", len(O)), ("Coagulation measurements", len(D)),
                     ("Per-position modification records", len(M)),
                     ("Oligonucleotides with per-position chemistry", len({r["oligo_id"] for r in M})),
                     ("Distinct sources", len(S))])
    block("Coverage", [("Oligonucleotides with a published sequence", len(seq)),
                       ("Oligonucleotides with a reported purity value", sum(1 for r in O if r["purity_pct"] not in (NR, "NOT_APPLICABLE", ""))),
                       ("Oligonucleotides with BOTH human and animal data", sum(1 for r in O if r["has_human_and_animal_data"] == "TRUE"))])
    block("Human versus animal (species_class)", [(k, v) for k, v in sc.most_common()])
    block("Study type", [(k, v) for k, v in st.most_common()])
    block("Measurements in a human or human-derived system", [("human_system = TRUE", sum(1 for r in D if r["human_system"] == "TRUE"))])
    block("Readout category", [(k, v) for k, v in rc.most_common()])
    block("Severity grade (provisional)", [
        ("grade 0 - no coagulation signal", gr.get("0", 0)),
        ("grade 1 - mild", gr.get("1", 0)),
        ("grade 2 - moderate", gr.get("2", 0)),
        ("grade 3 - severe", gr.get("3", 0)),
        ("ungraded (no published criterion applies; reason in grade_basis)", gr.get(NR, 0)),
        ("of the graded rows, flagged within_reference_range_resolution", sum(1 for r in D if r["grade_caveat"] == "within_reference_range_resolution"))])
    block("Axis", [("on-target pharmacology only", sum(1 for r in D if r["on_target_effect"] == "TRUE" and r["unintended_toxicity"] == "FALSE")),
                   ("unintended toxicity only", sum(1 for r in D if r["on_target_effect"] == "FALSE" and r["unintended_toxicity"] == "TRUE")),
                   ("both", sum(1 for r in D if r["on_target_effect"] == "TRUE" and r["unintended_toxicity"] == "TRUE")),
                   ("neither (context rows)", sum(1 for r in D if r["on_target_effect"] == "FALSE" and r["unintended_toxicity"] == "FALSE"))])
    hum_rows = [r for r in D if r["species_class"] == "human"]
    hum_oligos = {r["oligo_id"] for r in hum_rows}
    block("The human subset (see the human_measurements sheet)", [
        ("Human measurements", len(hum_rows)),
        ("Compounds measured in humans", len(hum_oligos)),
        ("…of which a published sequence is held", sum(1 for i in hum_oligos
            if next(r for r in O if r["oligo_id"] == i)["sequence_base"] not in (NR, "NOT_APPLICABLE", ""))),
        ("…of which per-position chemistry is held", len(hum_oligos & {r["oligo_id"] for r in M})),
        ("Human rows carrying a sequence", sum(1 for r in hum_rows
            if next(o for o in O if o["oligo_id"] == r["oligo_id"])["sequence_base"] not in (NR, "NOT_APPLICABLE", ""))),
        ("Human rows carrying an assigned grade", sum(1 for r in hum_rows if r["coag_tox_grade"] != NR)),
    ])
    block("Redistribution", [(k, v) for k, v in Counter(r["redistribution"] for r in D).most_common()])
    ws.column_dimensions["A"].width = 66
    ws.column_dimensions["B"].width = 16

    # ---- data_dictionary -------------------------------------------------
    dd = []
    for tbl, rows in (("sources", S), ("oligos", O), ("measurements", D), ("modifications", M)):
        for c in (rows[0].keys() if rows else []):
            dd.append({"table": tbl, "column": c, "definition": DEFS.get((tbl, c), DEFS.get(("*", c), "see schema.md"))})
    sheet(wb, "data_dictionary", dd)

    sheet(wb, "sources", S)
    sheet(wb, "oligos", O)
    sheet(wb, "measurements", D)
    sheet(wb, "modifications", M)


    # ---- human_measurements ------------------------------------------------
    # Every human row, with the compound's sequence and the toxicity score carried onto it,
    # so the most important subset of the dataset is usable without a join.
    mm, tx = modification_maps(M), toxicity_rollup(D)
    om = {r["oligo_id"]: r for r in O}
    OLIGO_COLS = ["oligo_name", "aliases", "oligo_class", "target_gene", "indication",
                  "developer", "max_phase", "length_nt", "sequence_5to3_asprinted",
                  "sequence_base", "sequence_note", "terminal_modification", "sequence_locus",
                  "backbone_chemistry", "sugar_modifications", "gapmer_design", "conjugate",
                  "ps_count", "purity_pct", "purity_method", "identity_confirmation"]
    hum = []
    for r in D:
        if r["species_class"] != "human":
            continue
        o = om[r["oligo_id"]]
        row = {"measurement_id": r["measurement_id"], "oligo_id": r["oligo_id"]}
        row.update({c: o[c] for c in OLIGO_COLS})
        m = mm.get(r["oligo_id"], {})
        row["sugar_modification_map"] = m.get("sugar_map", "NOT_REPORTED")
        row["backbone_linkage_map"] = m.get("backbone_map", "NOT_REPORTED")
        row["n_modification_positions"] = m.get("n_positions", 0)
        row["sequence_status"] = sequence_status(o)
        row.update({c: r[c] for c in D[0] if c not in ("measurement_id", "oligo_id")})
        hum.append(row)
    sheet(wb, "human_measurements", hum)

    # ---- German's analysis -------------------------------------------------
    # One row per compound: the oligo, its sequence, the modification to that sequence, and
    # its toxicity. Nothing else.
    ga = []
    for o in sorted(O, key=lambda r: r["oligo_id"]):
        m, t = mm.get(o["oligo_id"], {}), tx.get(o["oligo_id"], {})
        ga.append({
            "oligo_id": o["oligo_id"], "oligo_name": o["oligo_name"], "aliases": o["aliases"],
            "oligo_class": o["oligo_class"], "target_gene": o["target_gene"],
            "max_phase": o["max_phase"], "developer": o["developer"],
            "length_nt": o["length_nt"],
            "sequence_5to3_asprinted": o["sequence_5to3_asprinted"],
            "sequence_base": o["sequence_base"],
            "sequence_note": o["sequence_note"],
            "sequence_status": sequence_status(o),
            "terminal_modification": o["terminal_modification"],
            "backbone_chemistry": o["backbone_chemistry"],
            "sugar_modifications": o["sugar_modifications"],
            "gapmer_design": o["gapmer_design"], "conjugate": o["conjugate"], "ps_count": o["ps_count"],
            "sugar_modification_map": m.get("sugar_map", "NOT_REPORTED"),
            "backbone_linkage_map": m.get("backbone_map", "NOT_REPORTED"),
            "per_position_chemistry": m.get("per_position", "NOT_REPORTED"),
            "n_modification_positions": m.get("n_positions", 0),
            "worst_coag_tox_grade": t.get("worst_coag_tox_grade", "NOT_REPORTED"),
            "grade_distribution_0_1_2_3": t.get("grade_distribution_0_1_2_3", "NOT_REPORTED"),
            "n_graded_rows": t.get("n_graded_rows", 0),
            "n_rows_flagged_grade_caveat": t.get("n_rows_flagged_grade_caveat", 0),
            "max_source_stated_grade": t.get("max_source_stated_grade", "NOT_APPLICABLE"),
            "severity_in_source_words": t.get("severity_in_source_words", "NOT_REPORTED"),
            "n_measurements": o["n_measurements"],
            "n_human_rows": t.get("n_human_rows", 0), "n_animal_rows": t.get("n_animal_rows", 0),
            "on_target_rows": t.get("on_target_rows", 0),
            "unintended_toxicity_rows": t.get("unintended_toxicity_rows", 0),
            "has_human_and_animal_data": o["has_human_and_animal_data"],
            "source_ids": o["source_ids"],
        })
    sheet(wb, "German's analysis", ga)

    wb.save(OUT)
    print(f"  wrote {os.path.relpath(OUT, ROOT)}")
    print(f"    {len(S)} sources · {len(O)} oligos · {len(D)} measurements · {len(M)} modification rows")
    print(f"    data_dictionary: {len(dd)} columns documented")
    seq_ok = sum(1 for r in hum if r["sequence_base"] not in ("NOT_REPORTED", "NOT_APPLICABLE", ""))
    gr_ok = sum(1 for r in hum if r["coag_tox_grade"] != "NOT_REPORTED")
    print(f"    human_measurements: {len(hum)} rows — {seq_ok} carry a sequence, {gr_ok} carry a grade")
    print(f"    German's analysis:  {len(ga)} compounds — "
          f"{sum(1 for r in ga if r['sequence_base'] not in ('NOT_REPORTED','NOT_APPLICABLE',''))} with a sequence, "
          f"{sum(1 for r in ga if r['n_modification_positions'])} with per-position chemistry")


DEFS = {
    ("*", "oligo_id"): "Foreign key to oligos.oligo_id.",
    ("oligos", "oligo_id"): "Primary key. Stable identifier, COG-OLGnnn.",
    ("oligos", "oligo_name"): "Common or development name.",
    ("oligos", "aliases"): "Other names, semicolon separated.",
    ("oligos", "oligo_class"): "ASO_gapmer | ASO_mixmer | splice_switching_ASO | siRNA | GalNAc_siRNA | aptamer | PMO | tcDNA_ASO | CpG_ODN | polydisperse_ssDNA | other",
    ("oligos", "modality"): "single_stranded_ASO | double_stranded_siRNA | aptamer | mixture | other",
    ("oligos", "target_gene"): "Intended molecular target.",
    ("oligos", "indication"): "Disease or research context.",
    ("oligos", "developer"): "Originating organisation.",
    ("oligos", "max_phase"): "Highest development phase reached.",
    ("oligos", "length_nt"): "Length in nucleotides AS DECLARED BY THE SOURCE.",
    ("oligos", "length_nt_from_sequence"): "Length COMPUTED from sequence_base. Held separately so the declared value can be checked rather than trusted.",
    ("oligos", "sequence_5to3_asprinted"): "Sequence exactly as printed by the source, preserving any case convention that encodes chemistry.",
    ("oligos", "sequence_base"): "Nucleobases only, upper case, chemistry stripped. NOT_APPLICABLE for duplexes and polydisperse mixtures.",
    ("oligos", "sequence_note"): "Anything the source said about the sequence or length that is not itself sequence.",
    ("oligos", "terminal_modification"): "A terminal residue with no position in a 5'->3' string, e.g. a 3'-inverted dT cap.",
    ("oligos", "sequence_locus"): "Where in the document the sequence is printed.",
    ("oligos", "backbone_chemistry"): "full_PS | mixed_PO_PS | full_PO | PMO_neutral | other | NOT_REPORTED",
    ("oligos", "sugar_modifications"): "Semicolon-separated sugar chemistries.",
    ("oligos", "gapmer_design"): "Wing-gap-wing motif where applicable.",
    ("oligos", "conjugate"): "Conjugate moiety (GalNAc, PEG, cholesterol, none).",
    ("oligos", "ps_count"): "Number of phosphorothioate linkages.",
    ("oligos", "purity_pct"): "Per-compound purity. NOT_REPORTED throughout this release - see METHODOLOGY section 6.",
    ("oligos", "purity_method"): "Purification method as the source states it.",
    ("oligos", "identity_confirmation"): "Identity-confirmation method as the source states it.",
    ("oligos", "synthesis_platform"): "Synthesis platform as the source states it.",
    ("oligos", "source_ids"): "Semicolon-separated sources describing this compound.",
    ("oligos", "n_measurements"): "Measurement rows for this compound.",
    ("oligos", "n_human_measurements"): "Measurement rows in a human or human-derived system.",
    ("oligos", "n_animal_measurements"): "Measurement rows in an animal system.",
    ("oligos", "has_human_and_animal_data"): "TRUE where the compound is measured in BOTH - a human/animal translation pair.",
    ("oligos", "notes"): "Free text.",
    ("sources", "source_id"): "Primary key, COG-Snnn.",
    ("sources", "citation"): "Full citation as the document states it.",
    ("sources", "identifier"): "PMCID / PMID / DOI / US patent number / DailyMed set id.",
    ("sources", "document_file"): "File in sources/documents/ - the row's evidence is re-readable from the release.",
    ("sources", "retrieval_route"): "How the document was obtained.",
    ("sources", "licence"): "Licence as stated by the source.",
    ("sources", "redistribution"): "public_domain | CC_BY | CC_BY_NC | CC_BY_NC_ND | publisher_restricted | unresolved",
    ("sources", "extraction_bundle"): "Which extraction bundle read this source (audit trail).",
    ("sources", "n_oligos"): "Compounds described by this source.",
    ("sources", "n_measurements"): "Measurement rows from this source.",
    ("measurements", "measurement_id"): "Primary key, COG-MSRnnnn.",
    ("measurements", "source_id"): "Foreign key to sources.source_id.",
    ("measurements", "study_type"): "in_vitro | ex_vivo_plasma | animal_invivo | clinical. Carries the DESIGN only; species is carried by species_class.",
    ("measurements", "species"): "Species as the source states it, or NOT_APPLICABLE for a purified system.",
    ("measurements", "species_class"): "human | animal | not_determined. The human-versus-animal axis. A purified-protein assay is human when the proteins are human.",
    ("measurements", "species_class_basis"): "How species_class was determined: the species field, the system description, or source verification.",
    ("measurements", "human_system"): "TRUE where the measurement is made in a human or human-derived system - the Challenge's 'in vitro human system' criterion.",
    ("measurements", "system_model"): "Cell line, model, subject or assay system.",
    ("measurements", "matrix"): "plasma | whole_blood | serum | in_vivo | purified_system | NOT_APPLICABLE",
    ("measurements", "delivery_method"): "Route or delivery mode.",
    ("measurements", "dose_value"): "Dose or concentration.",
    ("measurements", "dose_unit"): "Unit of dose_value.",
    ("measurements", "timepoint"): "Timepoint of the measurement.",
    ("measurements", "exposure_duration"): "Duration of exposure.",
    ("measurements", "n_subjects"): "Subjects or animals contributing. Several clinical rows are underpowered; this makes that visible.",
    ("measurements", "is_baseline"): "TRUE for a pre-dose draw. A baseline is a reference point, not an effect: it carries no grade.",
    ("measurements", "co_administered_agent"): "Partner drug in a combination arm. A row with this set is NOT a measurement of the oligonucleotide alone.",
    ("measurements", "readout_category"): "clotting_time | factor_activity | fibrinogen | thrombin_generation | fibrinolysis_marker | anticoagulant_activity | bleeding_outcome | thrombotic_outcome | platelet_coag_crosstalk",
    ("measurements", "readout_name"): "The specific assay, e.g. aPTT, PT, INR, TT, ACT, fibrinogen, D_dimer, anti_Xa, FXI_activity, antithrombin_activity.",
    ("measurements", "readout_value"): "The value EXACTLY as printed, including any +/- or range.",
    ("measurements", "readout_unit"): "Unit of readout_value.",
    ("measurements", "readout_is_qualitative"): "TRUE where the row carries no number (e.g. the value exists only in a figure panel).",
    ("measurements", "control_value"): "The matched control value as printed.",
    ("measurements", "control_description"): "What the control was.",
    ("measurements", "effect_direction"): "increase | decrease | no_change | NOT_REPORTED | NOT_APPLICABLE. no_change means a MEASURED null, never an unmentioned endpoint.",
    ("measurements", "effect_vs_control"): "Effect size as the source expresses it.",
    ("measurements", "ratio_to_control"): "Control-referenced ratio computed at build time.",
    ("measurements", "ratio_basis"): "How the ratio was derived, or why none could be.",
    ("measurements", "coag_tox_grade"): "Ordinal 0-3 by CTCAE v5.0 cut-offs, or NOT_REPORTED where no published criterion applies.",
    ("measurements", "grade_basis"): "The exact rule applied, or for an ungraded row why no rule applies.",
    ("measurements", "grade_status"): "provisional on every row - no subject-matter review has taken place.",
    ("measurements", "grade_caveat"): "within_reference_range_resolution where the grade rests on a 1.0-1.2x ratio, which normal variation cannot be excluded from.",
    ("measurements", "source_stated_grade"): "The severity grade the SOURCE itself reports. A different rule from coag_tox_grade; a severity query should read both.",
    ("measurements", "severity_stated_by_source"): "Severity in the source's own words, verbatim.",
    ("measurements", "on_target_effect"): "TRUE where the compound is DESIGNED to act on coagulation.",
    ("measurements", "unintended_toxicity"): "TRUE where the source presents the finding as an adverse or unintended effect. Both flags may be true.",
    ("*", "sequence_status"): "Why a compound has no sequence: published | not_applicable_polydisperse_mixture | not_applicable_duplex_two_strands | recoverable_from_WHO_INN_nomenclature | not_published_by_source.",
    ("*", "sugar_modification_map"): "Per-position sugar chemistry, run-length encoded, e.g. 1-5:2'-MOE; 6-15:DNA; 16-20:2'-MOE.",
    ("*", "backbone_linkage_map"): "Per-position backbone linkage, run-length encoded.",
    ("*", "per_position_chemistry"): "The full per-residue map: position, nucleobase, sugar and 3' linkage.",
    ("*", "worst_coag_tox_grade"): "Highest ASSIGNED grade across the compound's rows. NOT_REPORTED where no row could be graded -- absence of a grade is not a grade of 0.",
    ("*", "grade_distribution_0_1_2_3"): "Count of the compound's rows at each grade.",
    ("*", "severity_in_source_words"): "Severity as the source states it, verbatim.",
    ("measurements", "value_origin"): "measured_in_this_document | cited_from_another_source",
    ("measurements", "source_locus"): "Exact locus - table, figure, section, label section or page.",
    ("measurements", "redistribution"): "Inherited from the source.",
    ("measurements", "verbatim_quote"): "Text copied from the document supporting this row. Present on every row.",
    ("measurements", "notes"): "Free text, including method limitations and reporting-silence flags.",
    ("modifications", "position"): "Nucleotide position, 5' to 3', contiguous from 1.",
    ("modifications", "nucleobase"): "A | C | G | T | U. Must equal sequence_base at this position.",
    ("modifications", "sugar_mod"): "DNA | RNA | LNA | 2'-MOE | 2'-OMe | 2'-F | cEt | morpholino | tcDNA | NOT_REPORTED",
    ("modifications", "backbone_linkage_3p"): "The linkage 3' of this position: PS | PO | PN | NOT_APPLICABLE | NOT_REPORTED",
    ("modifications", "is_5_methyl_C"): "TRUE only where the source states it. FALSE means not stated.",
    ("modifications", "basis"): "How the chemistry at this position was determined, e.g. the source's own case legend quoted.",
}

if __name__ == "__main__":
    main()
