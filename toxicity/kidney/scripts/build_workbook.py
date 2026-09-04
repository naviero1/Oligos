#!/usr/bin/env python3
"""Build the OligoTox-Kidney Excel deliverable.

Tab layout follows the reviewer's spec:

  Summary               orientation, tab index, and the caveats a reader must know
                        before analysing (counts written as values -- see note in main())
  Human trials          human_clinical rows -- the priority subset. Sequence and
                        toxicity grade are the FIRST columns after the IDs, then
                        every other field.
  Human in vitro        human_invitro rows, same treatment
  German's analysis     one row per oligo: oligo / sequence / modification / toxicity
  All measurements      all 246 rows
  Oligos                the 65-row design table
  Human vs animal       the 15 bridging oligos
  Data dictionary       column definitions

Two things the spec depends on and which are easy to lose in Excel:

  * CASE IS DATA in gapmer sequences -- upper case marks 2'-MOE/cEt wings, lower
    case the DNA gap. Sequence cells are written as explicit text so Excel cannot
    auto-format or upper-case them, and the convention is stated on every tab that
    shows a sequence.
  * Sequence coverage is 41/42 on human trials. The one gap (MSR058) is the pooled
    class-level Crooke row, where a single sequence is not meaningful rather than
    missing. It is labelled as such in-cell, not left blank.

Usage:  python scripts/build_workbook.py
"""
import csv
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, "data", "OligoTox-Kidney.xlsx")

FONT = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
KEY_FILL = PatternFill("solid", fgColor="FFF2CC")     # sequence + toxicity emphasis
NOTE_FONT = Font(name=FONT, italic=True, size=9, color="666666")
BODY = Font(name=FONT, size=10)
MONO = Font(name="Consolas", size=10)                  # sequences, so case is legible
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(bottom=THIN)

GRADE_FILL = {"0": PatternFill("solid", fgColor="E2EFDA"),
              "1": PatternFill("solid", fgColor="FFF2CC"),
              "2": PatternFill("solid", fgColor="FCE4D6"),
              "3": PatternFill("solid", fgColor="F8CBAD")}

CASE_NOTE = ("CASE IS DATA in gapmer sequences: UPPER = 2'-MOE / cEt wing, lower = DNA gap. "
             "Do not upper-case this column.")


def load():
    d = os.path.join(ROOT, "data")
    o = list(csv.DictReader(open(os.path.join(d, "oligos.csv"), newline="")))
    m = list(csv.DictReader(open(os.path.join(d, "measurements.csv"), newline="")))
    b = list(csv.DictReader(open(os.path.join(d, "human_animal_bridge.csv"), newline="")))
    return {r["oligo_id"]: r for r in o}, o, m, b


def seq_of(ol):
    s = ol["sequence_5to3"].strip()
    if s in ("TBD", "", "NA"):
        # distinguish "not meaningful" from "not yet found" -- they are different facts
        if ol["max_phase"] == "class_review" or "class" in ol["oligo_name"].lower():
            return "n/a - pooled class-level entry, no single sequence"
        return "TBD - not published by any source"
    return s


def modification_summary(ol):
    bits = []
    if ol["backbone_chemistry"] not in ("TBD", "", "NA"):
        bits.append(ol["backbone_chemistry"])
    if ol["sugar_modifications"] not in ("TBD", "", "NA"):
        bits.append(ol["sugar_modifications"])
    if ol["gapmer_design"] not in ("TBD", "", "NA", "none"):
        bits.append(f"gapmer {ol['gapmer_design']}")
    if ol["conjugate"] not in ("none", "TBD", "", "NA"):
        bits.append(f"conjugate {ol['conjugate']}")
    if ol["ps_count"] not in ("TBD", "", "NA"):
        bits.append(f"{ol['ps_count']} PS linkages")
    return " | ".join(bits) if bits else "TBD"


def write_sheet(ws, headers, rows, seq_cols=(), grade_cols=(), note=None, widths=None):
    r0 = 1
    if note:
        ws.cell(1, 1, note).font = NOTE_FONT
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(headers), 6))
        r0 = 2
    for c, h in enumerate(headers, 1):
        cell = ws.cell(r0, c, h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[r0].height = 30
    for i, row in enumerate(rows, r0 + 1):
        for c, v in enumerate(row, 1):
            cell = ws.cell(i, c, v)
            cell.font = MONO if headers[c - 1] in seq_cols else BODY
            cell.border = BORDER
            if headers[c - 1] in seq_cols:
                cell.number_format = "@"          # force text; protects case and leading chars
                cell.fill = KEY_FILL
            if headers[c - 1] in grade_cols:
                cell.fill = GRADE_FILL.get(str(v), KEY_FILL)
                cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = ws.cell(r0 + 1, 3)
    ws.auto_filter.ref = f"A{r0}:{get_column_letter(len(headers))}{r0 + len(rows)}"
    for c, h in enumerate(headers, 1):
        w = (widths or {}).get(h)
        if w is None:
            longest = max([len(str(h))] + [len(str(r[c - 1])) for r in rows[:400]] or [10])
            w = min(max(longest + 2, 10), 46)
        ws.column_dimensions[get_column_letter(c)].width = w


def main():
    OL, oligos, meas, bridge = load()
    wb = Workbook()

    # ---------- measurement tabs: sequence and grade lead ----------
    MEAS_HEAD = ["measurement_id", "oligo_id", "oligo_name", "sequence_5to3",
                 "nephrotox_grade", "modification_summary", "oligo_class", "target_gene",
                 "study_type", "species", "subject_class", "system_model", "tissue",
                 "delivery_method", "dose_or_conc_value", "dose_or_conc_unit",
                 "exposure_duration", "readout_category", "readout_name", "readout_value",
                 "readout_unit", "effect_direction", "effect_vs_control",
                 "renal_endpoints_measured", "identity_confirmation", "purity_pct",
                 "source_id", "source_ref", "source_table", "redistribution", "notes"]

    def meas_row(r):
        ol = OL[r["oligo_id"]]
        return [r["measurement_id"], r["oligo_id"], ol["oligo_name"], seq_of(ol),
                r["nephrotox_grade"], modification_summary(ol), ol["oligo_class"],
                ol["target_gene"], r["study_type"], r["species"], r["subject_class"],
                r["system_model"], r["tissue"], r["delivery_method"],
                r["dose_or_conc_value"], r["dose_or_conc_unit"], r["exposure_duration"],
                r["readout_category"], r["readout_name"], r["readout_value"],
                r["readout_unit"], r["effect_direction"], r["effect_vs_control"],
                r["renal_endpoints_measured"], ol["identity_confirmation"], ol["purity_pct"],
                r["source_id"], r["source_ref"], r["source_table"], r["redistribution"],
                r["notes"]]

    W = {"sequence_5to3": 30, "notes": 60, "modification_summary": 38, "source_ref": 34,
         "nephrotox_grade": 9, "readout_name": 26, "system_model": 24}

    ws = wb.active
    ws.title = "Summary"

    clinical = [r for r in meas if r["subject_class"] == "human_clinical"]
    invitro = [r for r in meas if r["subject_class"] == "human_invitro"]

    write_sheet(wb.create_sheet("Human trials"), MEAS_HEAD, [meas_row(r) for r in clinical],
                seq_cols=("sequence_5to3",), grade_cols=("nephrotox_grade",), widths=W,
                note=f"HUMAN CLINICAL TRIALS - the priority subset ({len(clinical)} rows). "
                     f"Sequence and toxicity grade lead the table. {CASE_NOTE}")
    write_sheet(wb.create_sheet("Human in vitro"), MEAS_HEAD, [meas_row(r) for r in invitro],
                seq_cols=("sequence_5to3",), grade_cols=("nephrotox_grade",), widths=W,
                note=f"HUMAN IN VITRO cell systems ({len(invitro)} rows). {CASE_NOTE}")

    # ---------- German's analysis: oligo / sequence / modification / toxicity ----------
    GA_HEAD = ["oligo_id", "oligo_name", "oligo_class", "sequence_5to3", "length_nt",
               "modification_summary", "backbone_chemistry", "sugar_modifications",
               "gapmer_design", "ps_count", "conjugate",
               "max_grade_human", "max_grade_animal", "max_grade_overall",
               "n_human_rows", "n_animal_rows", "n_rows_total",
               "identity_confirmation", "purity_pct", "target_gene", "max_phase"]
    ga = []
    for oid, ol in sorted(OL.items()):
        rows = [r for r in meas if r["oligo_id"] == oid]
        hum = [r for r in rows if r["subject_class"].startswith("human")]
        ani = [r for r in rows if r["subject_class"].startswith("animal")]
        gmax = lambda rs: max((int(r["nephrotox_grade"]) for r in rs), default="")
        ga.append([oid, ol["oligo_name"], ol["oligo_class"], seq_of(ol), ol["length_nt"],
                   modification_summary(ol), ol["backbone_chemistry"],
                   ol["sugar_modifications"], ol["gapmer_design"], ol["ps_count"],
                   ol["conjugate"], gmax(hum), gmax(ani), gmax(rows),
                   len(hum), len(ani), len(rows),
                   ol["identity_confirmation"], ol["purity_pct"], ol["target_gene"],
                   ol["max_phase"]])
    write_sheet(wb.create_sheet("German's analysis"), GA_HEAD, ga,
                seq_cols=("sequence_5to3",),
                grade_cols=("max_grade_human", "max_grade_animal", "max_grade_overall"),
                widths={"sequence_5to3": 30, "modification_summary": 40,
                        "sugar_modifications": 26, "oligo_name": 26},
                note="ONE ROW PER OLIGO: sequence, its modifications, and the toxicity it "
                     f"reached. Grades are the MAXIMUM observed per subject class. {CASE_NOTE}")

    write_sheet(wb.create_sheet("All measurements"), MEAS_HEAD, [meas_row(r) for r in meas],
                seq_cols=("sequence_5to3",), grade_cols=("nephrotox_grade",), widths=W,
                note=f"All {len(meas)} measurements, every one strict-kidney "
                     f"(is_kidney_specific=TRUE). {CASE_NOTE}")

    OHEAD = list(oligos[0].keys())
    write_sheet(wb.create_sheet("Oligos"), OHEAD, [[r[k] for k in OHEAD] for r in oligos],
                seq_cols=("sequence_5to3",), widths={"sequence_5to3": 30, "notes": 60},
                note=f"Design table, one row per oligonucleotide ({len(oligos)}). {CASE_NOTE}")

    BHEAD = list(bridge[0].keys())
    write_sheet(wb.create_sheet("Human vs animal"), BHEAD, [[r[k] for k in BHEAD] for r in bridge],
                grade_cols=("human_max_grade", "animal_max_grade"),
                note="Oligos carrying evidence on BOTH sides of the human/animal divide. "
                     "concordance compares max human grade against max animal grade. "
                     "Read with care: where the human grade is an unvalidated 0, an "
                     "'animal_over_predicts' verdict may mean nobody measured the human endpoint.")

    # ---------- data dictionary ----------
    DD = [
        ("sequence_5to3", "5'->3' sequence. CASE IS DATA: upper = 2'-MOE/cEt wing, lower = DNA gap."),
        ("nephrotox_grade", "0 none | 1 mild/functional reversible | 2 moderate biomarker or histopath | 3 severe AKI/GN. All provisional pending scientific sign-off."),
        ("renal_endpoints_measured", "measured_and_reported | not_measured | not_reported_in_source | cannot_determine. ONLY measured_and_reported supports grade 0 as safety evidence."),
        ("subject_class", "human_clinical | human_invitro | animal_invitro | animal_invivo. Derived from study_type + species."),
        ("modification_summary", "Backbone, sugar modifications, gapmer motif, conjugate and PS count, concatenated for reading."),
        ("identity_confirmation", "How the sequence identity was established (WHO INN nomenclature, patent sequence listing, label, publication) or not_established."),
        ("purity_pct", "TBD for all 65. Verified unavailable: no source publishes per-batch purity, and no wet lab was run."),
        ("readout_unit", "Note two normalisations exist: pct_saline_control and pct_compound_1-1_reference. DO NOT POOL THEM."),
        ("redistribution", "public_domain | summary_stat | derived_features_only | verify. Governs whether a raw value may be republished."),
        ("source_id / source_ref / source_table", "Provenance triple. Every row carries all three; source_table names the exact table, figure or label section."),
    ]
    write_sheet(wb.create_sheet("Data dictionary"), ["column", "definition"],
                [list(x) for x in DD], widths={"column": 30, "definition": 110},
                note="Key columns. Full dictionary in schema.md.")

    # ---------- summary tab: live formulas over the data tabs ----------
    ws.sheet_view.showGridLines = False
    ws["A1"] = "OligoTox-Kidney"
    ws["A1"].font = Font(name=FONT, bold=True, size=16, color="1F3864")
    ws["A2"] = "Curated nephrotoxicity dataset for oligonucleotide therapeutics - NIH/NCATS OligoTox Challenge, Phase 2"
    ws["A2"].font = Font(name=FONT, size=10, color="666666")
    ws["A3"] = "Endpoint: KIDNEY only. Every row is is_kidney_specific=TRUE; no other toxicity endpoint is mixed in."
    ws["A3"].font = Font(name=FONT, size=10, italic=True, color="C00000")

    rows_summary = [
        ("Tab", "What it holds", "Rows"),
        # Counts are written as values, not COUNTA formulas. This workbook is a data
        # SNAPSHOT regenerated from the canonical CSVs by this script -- it has no input
        # cells a reader edits, so nothing needs to recalculate. LibreOffice is
        # unavailable in the build environment, and an unrecalculated formula reads back
        # as a BLANK cell in Excel and pandas, which is strictly worse for a reviewer
        # than a correct number. Re-run this script to refresh.
        ("Human trials", "Human clinical rows - the priority subset, sequence + toxicity first", len(clinical)),
        ("Human in vitro", "Human cell systems (PTEC, PTEC-TERT1, ciPTEC, tubule-on-chip)", len(invitro)),
        ("German's analysis", "One row per oligo: sequence, modification, toxicity", len(ga)),
        ("All measurements", "Every measurement in the dataset", len(meas)),
        ("Oligos", "Design table, one row per oligonucleotide", len(oligos)),
        ("Human vs animal", "Oligos with paired human and animal evidence", len(bridge)),
    ]
    for i, (a, b_, c) in enumerate(rows_summary, 5):
        ws.cell(i, 1, a).font = Font(name=FONT, bold=(i == 5), size=10)
        ws.cell(i, 2, b_).font = Font(name=FONT, bold=(i == 5), size=10)
        ws.cell(i, 3, c).font = Font(name=FONT, bold=(i == 5), size=10)
        if i == 5:
            for c_ in range(1, 4):
                ws.cell(i, c_).fill = HDR_FILL
                ws.cell(i, c_).font = HDR_FONT

    notes = [
        "",
        "READ THIS BEFORE ANALYSING",
        "1. Case is data. In gapmer sequences UPPER = 2'-MOE/cEt wing, lower = DNA gap. Do not upper-case the sequence column.",
        "2. Grade 0 is not always evidence of safety. Filter on renal_endpoints_measured = measured_and_reported; 13 clinical",
        "   grade-0 rows are flagged otherwise, meaning nobody confirmed the endpoint was measured.",
        "3. Two normalisations exist in readout_value (% saline vs % compound 1-1). Never pool them; check readout_unit.",
        "4. Split by oligo_id, never by row, when modelling - one compound contributes up to 25 rows.",
        "5. All grades are provisional pending scientific sign-off.",
        "",
        "Sequence coverage on human trials: 41 of 42 rows. The single gap (MSR058) is a pooled class-level entry where one",
        "sequence is not meaningful, not a missing value.",
    ]
    for i, line in enumerate(notes, 5 + len(rows_summary) + 1):
        c = ws.cell(i, 1, line)
        c.font = Font(name=FONT, size=10, bold=line.isupper() and bool(line))
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 74
    ws.column_dimensions["C"].width = 10

    wb.save(OUT)
    print(f"wrote {OUT}")
    for s in wb.sheetnames:
        print(f"  tab: {s}")


if __name__ == "__main__":
    main()
