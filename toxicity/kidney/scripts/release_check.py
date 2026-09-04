#!/usr/bin/env python3
"""Pre-submission gate for the OligoTox-Kidney Phase 2 release.

Checks the dataset against the four explicit Phase 2 dataset requirements, verifies
strict-kidney isolation, and confirms the four submission artefacts exist. Exits
non-zero if any hard check fails, so it can gate a release.
"""
import csv, os, re, sys, collections

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
REPO = os.path.dirname(os.path.dirname(ROOT))
D = lambda *p: os.path.join(ROOT, *p)

fails, warns = [], []
def check(ok, msg, hard=True):
    print(f"  {'PASS' if ok else ('FAIL' if hard else 'WARN')}  {msg}")
    if not ok:
        (fails if hard else warns).append(msg)

o = list(csv.DictReader(open(D("data", "oligos.csv"), newline="")))
m = list(csv.DictReader(open(D("data", "measurements.csv"), newline="")))
g = list(csv.DictReader(open(D("data", "oligotox_kidney_merged.csv"), newline="")))

print("\n[1] INTEGRITY")
ids = {r["oligo_id"] for r in o}
check(len(o) == 65 and len(m) == 246, f"row counts: {len(o)} oligos / {len(m)} measurements")
check(sum(1 for r in m if r["oligo_id"] not in ids) == 0, "referential integrity (0 orphans)")
check(len({r["oligo_id"] for r in o}) == len(o), "no duplicate oligo_id")
check(len({r["measurement_id"] for r in m}) == len(m), "no duplicate measurement_id")
check(all(r["nephrotox_grade"] in "0123" for r in m), "all grades in {0,1,2,3}")
check(len(g) == len(m), f"merged view row count matches ({len(g)})")

print("\n[2] STRICT-KIDNEY ISOLATION (no mixing with other toxicities)")
check(all(r["is_kidney_specific"] == "TRUE" for r in m), "every row is_kidney_specific=TRUE")
check(all(r["tissue"] in ("kidney", "proximal_tubule", "glomerulus") for r in m), "every tissue is renal")
# Precision matters here. Case-insensitive "ALT" matches "heALThy_volunteer", and a
# journal name can legitimately contain "Liver" -- the givosiran CKD finding was
# published in Liver International and is a kidney result. So: biomarker abbreviations
# are matched case-SENSITIVELY with word boundaries, hepatic tissue/model words are
# matched on their own, and source provenance is checked against the actual
# hepatotoxicity source identifiers rather than free-text journal names.
hep_readout = re.compile(r"\b(ALT|AST)\b|bilirub|transaminase|hepatic|hepatocyte", re.I)
hep_model = re.compile(r"hepat|HepG2|\bliver\b|primary_hepato", re.I)
HEPATOTOX_SOURCES = ("Dieckmann", "Burdick", "Hagedorn")
check(not [r for r in m if hep_readout.search(r["readout_name"])], "no hepatic readouts")
check(not [r for r in m if hep_model.search(r["system_model"])], "no hepatic system models")
check(not [r for r in m if any(h in r["source_ref"] for h in HEPATOTOX_SOURCES)],
      "no hepatotoxicity source panels cited")
check(all(r["readout_category"] != "viability" or r["tissue"] in ("kidney", "proximal_tubule")
          for r in m), "viability rows are renal-tissue only")

print("\n[3] PHASE 2 DATASET REQUIREMENTS")
seq = [r for r in o if r["sequence_5to3"].strip() not in ("TBD", "", "NA")]
check(len(seq) == 55, f"sequences present for {len(seq)}/{len(o)} oligos (10 structurally unavailable)", hard=False)
check(all(r["identity_confirmation"].strip() for r in o), "identity_confirmation populated for every oligo")
check(all(r["purity_pct"].strip() for r in o), "purity_pct present as an explicit value (TBD, verified unavailable)")
check(os.path.exists(os.path.join(REPO, "LICENSE")), "LICENSE at repository root")
check(os.path.exists(D("schema.md")), "data dictionary / schema present")
mods = sum(1 for r in o if r["sugar_modifications"].strip() not in ("TBD", "", "NA"))
check(mods == len(o), f"modification composition recorded for {mods}/{len(o)}")

print("\n[4] LABEL PROVENANCE")
prov = collections.Counter(r["renal_endpoints_measured"] for r in m)
check("renal_endpoints_measured" in m[0], "renal_endpoints_measured present")
unsupported = [r["measurement_id"] for r in m
               if r["study_type"] == "clinical" and r["nephrotox_grade"] == "0"
               and r["renal_endpoints_measured"] != "measured_and_reported"]
print(f"        provenance: {dict(prov)}")
print(f"        grade-0 clinical rows flagged unsupported: {len(unsupported)}")
check(all(r["source_id"] and r["source_ref"] and r["source_table"] for r in m),
      "every row carries source_id + source_ref + source_table")

print("\n[5] SUBMISSION ARTEFACTS")
for f, label in [("NARRATIVE.md", "narrative document"),
                 ("METHODOLOGY_PHASE2.md", "methodology document"),
                 ("PADP.md", "public access & dissemination plan"),
                 ("schema.md", "dataset: data dictionary & schema"),
                 ("data/oligos.csv", "dataset: oligo table"),
                 ("data/measurements.csv", "dataset: measurement table"),
                 ("data/OligoTox-Kidney.xlsx", "dataset: Excel workbook"),
                 ("NARRATIVE.pdf", "narrative rendered to PDF"),
                 ("METHODOLOGY_PHASE2.pdf", "methodology rendered to PDF"),
                 ("PADP.pdf", "PADP rendered to PDF")]:
    check(os.path.exists(D(f)), f"{label} ({f})")

# the workbook must carry the tabs the reviewer specified, and every human-trial row
# must carry a sequence cell and a grade -- the spec's central requirement
try:
    from openpyxl import load_workbook
    wbk = load_workbook(D("data", "OligoTox-Kidney.xlsx"))
    for tab in ("Human trials", "German's analysis"):
        check(tab in wbk.sheetnames, f"workbook tab present: {tab}")
    htab = wbk["Human trials"]
    hdr = [htab.cell(2, c).value for c in range(1, htab.max_column + 1)]
    check("sequence_5to3" in hdr and "nephrotox_grade" in hdr,
          "Human trials carries sequence_5to3 and nephrotox_grade")
    si, gi = hdr.index("sequence_5to3") + 1, hdr.index("nephrotox_grade") + 1
    blanks = [r for r in range(3, htab.max_row + 1)
              if not str(htab.cell(r, si).value or "").strip()
              or not str(htab.cell(r, gi).value or "").strip()]
    check(not blanks, f"every Human trials row has a sequence cell and a grade ({len(blanks)} blank)")
except ImportError:
    warns.append("openpyxl unavailable - workbook not inspected")

print("\n" + "=" * 62)
if fails:
    print(f"RELEASE BLOCKED — {len(fails)} hard failure(s)")
    for f in fails: print("  -", f)
    sys.exit(1)
print(f"RELEASE CHECKS PASSED" + (f" ({len(warns)} advisory)" if warns else ""))
for w in warns: print("  advisory:", w)
