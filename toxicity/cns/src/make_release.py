#!/usr/bin/env python3
"""Build the two release artefacts that are generated rather than written:

    deliverables/OligoTox-CNS_Dataset.xlsx   the dataset workbook
    docs/DATA_DICTIONARY.md                  the data dictionary

Both are produced from the canonical column definitions in src/assemble.py and from the CSVs
in data/, so neither can drift away from the released data.

The Summary sheet uses live formulas (COUNTA / COUNTIF) against the data sheets rather than
values computed here, so the workbook recomputes its own headline numbers when opened.
"""
from __future__ import annotations

import csv
import pathlib
import sys

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

sys.path.insert(0, str(pathlib.Path(__file__).resolve().parent))
from assemble import (MEASUREMENT_COLUMNS, MODIFICATION_COLUMNS,  # noqa: E402
                      OLIGO_COLUMNS, SOURCES)

ROOT = pathlib.Path(__file__).resolve().parent.parent
DATA = ROOT / "data"
OUT_XLSX = ROOT / "deliverables" / "OligoTox-CNS_Dataset.xlsx"
OUT_MD = ROOT / "docs" / "DATA_DICTIONARY.md"

FONT = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1F3B5C")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
BODY = Font(name=FONT, size=10)
TITLE = Font(name=FONT, bold=True, size=14, color="1F3B5C")
H2 = Font(name=FONT, bold=True, size=11, color="1F3B5C")
THIN = Side(style="thin", color="D9D9D9")

TITLE_TEXT = "OligoTox-CNS — CNS toxicity of oligonucleotide therapeutics"
VERSION = "v1.0"
LICENCE = "Creative Commons Attribution 4.0 International (CC BY 4.0)"

README_LINES = [
    (TITLE_TEXT, "title"),
    (f"Dataset release {VERSION}  ·  licence: {LICENCE}", "sub"),
    ("", ""),
    ("What this is", "h2"),
    ("An open, sequence-resolved dataset of central-nervous-system toxicity measurements for "
     "oligonucleotide therapeutics, assembled for the NIH/NCATS Oligonucleotide Toxicity Open "
     "Data Challenge, Phase 2 (Data Generation).", ""),
    ("Every row pairs an oligonucleotide's design — its sequence and the position of every "
     "chemical modification in it — with a measured CNS toxicity outcome, so that a model can "
     "be trained to predict the second from the first.", ""),
    ("", ""),
    ("Sheets", "h2"),
    ("Summary            headline counts, computed live from the data sheets", "mono"),
    ("data_dictionary    every column in every table, with its definition", "mono"),
    ("oligos             one row per oligonucleotide — the predictor variables", "mono"),
    ("measurements       one row per measured outcome — the response variables", "mono"),
    ("modifications      one row per NUCLEOTIDE POSITION — the per-position chemistry", "mono"),
    ("sources            the provenance registry: citation, licence, retrieval route", "mono"),
    ("", ""),
    ("How the tables join", "h2"),
    ("measurements.oligo_id   -> oligos.oligo_id      (many measurements per oligonucleotide)", "mono"),
    ("modifications.oligo_id  -> oligos.oligo_id      (one row per position, 5' to 3')", "mono"),
    ("oligos.source_id        -> sources.source_id", "mono"),
    ("", ""),
    ("Missing values — read this before analysing", "h2"),
    ("NOT_REPORTED    the source does not report this value. It has NOT been estimated, "
     "imputed or filled from background knowledge.", ""),
    ("NOT_APPLICABLE  the field has no meaning for this row (e.g. a gap length for a vehicle "
     "control).", ""),
    ("An empty cell   the field does not apply to this table's row type.", ""),
    ("", ""),
    ("In particular: purity_pct is NOT_REPORTED for every oligonucleotide in this release. "
     "Per-compound purity values are almost never published alongside toxicity results. What "
     "IS captured, where the source states it, is the purification and identity-confirmation "
     "METHOD (purity_method, identity_confirmation). See the methodology document.", ""),
    ("", ""),
    ("A note on the Summary sheet", "h2"),
    ("The Summary sheet holds live formulas (COUNTA / COUNTIF) over the data sheets rather than "
     "values written in, so it recomputes from the data whenever the workbook is opened. Excel "
     "and LibreOffice calculate on open; a tool that reads cached values WITHOUT calculating "
     "(e.g. pandas.read_excel) will see those cells as empty. That is expected - read the CSVs "
     "in data/ for programmatic access. Each formula was independently checked against the CSVs "
     "and every one evaluates to the value qc/validate_dataset.py reports.", ""),
    ("", ""),
    ("Grades are provisional", "h2"),
    ("cns_tox_grade is an ordinal 0-3 severity assigned by the rule recorded in that row's "
     "grade_basis. Grades derived from the mouse acute tolerability score use the cut-offs "
     "published by Hagedorn et al. (2022), not thresholds invented here. All grades ship with "
     "grade_status = provisional pending subject-matter-expert review.", ""),
    ("", ""),
    ("How to cite", "h2"),
    ("Cite this dataset AND the underlying primary sources listed on the 'sources' sheet. "
     "Source H1 is CC BY 4.0; sources K1, L1 and O1 are CC BY-NC; source C1 is a US Government "
     "work in the public domain. The redistribution column on each measurement row records "
     "which terms apply to that row.", ""),
]


def read(name: str) -> list[dict]:
    with (DATA / f"{name}.csv").open() as fh:
        return list(csv.DictReader(fh))


def data_sheet(wb, name: str, columns, rows: list[dict], numeric_cols=()):
    ws = wb.create_sheet(name)
    heads = [c for c, _ in columns]
    ws.append(heads)
    for cell in ws[1]:
        cell.font, cell.fill = HDR_FONT, HDR_FILL
        cell.alignment = Alignment(vertical="center", horizontal="left")
    for r in rows:
        ws.append([r.get(c, "") for c in heads])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY
    # numeric coercion where the column really is numeric
    for col in numeric_cols:
        if col not in heads:
            continue
        idx = heads.index(col) + 1
        for row in ws.iter_rows(min_row=2, min_col=idx, max_col=idx):
            v = row[0].value
            if isinstance(v, str) and v.replace(".", "", 1).replace("-", "", 1).isdigit():
                row[0].value = float(v) if "." in v else int(v)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(heads))}{ws.max_row}"
    for i, h in enumerate(heads, start=1):
        sample = [str(r.get(h, "")) for r in rows[:400]]
        width = max([len(h)] + [len(s) for s in sample]) + 2
        ws.column_dimensions[get_column_letter(i)].width = min(max(width, 10), 46)
    ws.sheet_view.showGridLines = False
    return ws


def main() -> int:
    OUT_XLSX.parent.mkdir(exist_ok=True)
    OUT_MD.parent.mkdir(exist_ok=True)
    oligos, meas, mods, srcs = (read("oligos"), read("measurements"),
                                read("modifications"), read("sources"))
    src_cols = [(k, "") for k in srcs[0]]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ---------------- README -------------------------------------------------------------
    ws = wb.create_sheet("README")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 118
    for text, kind in README_LINES:
        ws.append([text])
        c = ws.cell(row=ws.max_row, column=1)
        c.font = {"title": TITLE, "h2": H2,
                  "sub": Font(name=FONT, size=10, italic=True, color="52514E"),
                  "mono": Font(name="Consolas", size=9.5)}.get(kind, BODY)
        c.alignment = Alignment(wrap_text=True, vertical="top")

    # ---------------- Summary (live formulas) --------------------------------------------
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 56
    ws.column_dimensions["B"].width = 16
    ws.append([TITLE_TEXT]); ws["A1"].font = TITLE
    ws.append([f"Release {VERSION}. Every figure below is a formula over the data sheets."])
    ws["A2"].font = Font(name=FONT, size=10, italic=True, color="52514E")
    ws.append([])

    g_col = get_column_letter([c for c, _ in MEASUREMENT_COLUMNS].index("cns_tox_grade") + 1)
    h_col = get_column_letter([c for c, _ in MEASUREMENT_COLUMNS].index("is_human_system") + 1)
    s_col = get_column_letter([c for c, _ in MEASUREMENT_COLUMNS].index("study_type") + 1)
    b_col = get_column_letter([c for c, _ in OLIGO_COLUMNS].index("modification_position_basis") + 1)
    p_col = get_column_letter([c for c, _ in OLIGO_COLUMNS].index("purity_pct") + 1)

    # Bounded ranges, not whole-column references: whole-column COUNTIF over a workbook
    # containing a 32,569-row sheet takes LibreOffice more than nine minutes to recalculate.
    NO = len(oligos) + 200          # oligos data rows + headroom
    NM = len(meas) + 200            # measurements data rows + headroom
    ND = len(mods) + 200            # modifications data rows + headroom
    NS = len(srcs) + 20
    orng = lambda c: f"oligos!${c}$2:${c}${NO}"
    mrng = lambda c: f"measurements!${c}$2:${c}${NM}"
    seq_col = get_column_letter([c for c, _ in OLIGO_COLUMNS].index("sequence_base") + 1)

    rows = [
        ("Counts", None),
        ("Oligonucleotides", f'=COUNTA({orng("A")})'),
        ("CNS toxicity measurements", f'=COUNTA({mrng("A")})'),
        ("Per-position modification records", f'=COUNTA(modifications!$A$2:$A${ND})'),
        ("Distinct sources", f'=COUNTA(sources!$A$2:$A${NS})'),
        (None, None),
        ("Coverage", None),
        ("Oligonucleotides with a published sequence",
         f'=COUNTA({orng("A")})-COUNTIF({orng(seq_col)},"NOT_REPORTED")'),
        ("Oligonucleotides with position-resolved modifications",
         f'=COUNTIF({orng(b_col)},"position_resolved*")'),
        ("Oligonucleotides with a reported purity value",
         f'=COUNTA({orng("A")})-COUNTIF({orng(p_col)},"NOT_REPORTED")'),
        (None, None),
        ("Measurements by study type", None),
        ("in vitro", f'=COUNTIF({mrng(s_col)},"in_vitro")'),
        ("in vivo (animal)", f'=COUNTIF({mrng(s_col)},"animal_invivo")'),
        ("clinical (human)", f'=COUNTIF({mrng(s_col)},"clinical")'),
        ("measured in a human or human-derived system",
         f'=COUNTIF({mrng(h_col)},"TRUE")'),
        (None, None),
        ("Severity grade distribution", None),
        ("grade 0 — no observable CNS signal", f'=COUNTIF({mrng(g_col)},0)'),
        ("grade 1 — mild", f'=COUNTIF({mrng(g_col)},1)'),
        ("grade 2 — moderate", f'=COUNTIF({mrng(g_col)},2)'),
        ("grade 3 — severe", f'=COUNTIF({mrng(g_col)},3)'),
    ]
    for label, formula in rows:
        if label is None:
            ws.append([])
            continue
        if formula is None:
            ws.append([label])
            ws.cell(row=ws.max_row, column=1).font = H2
        else:
            ws.append([label, formula])
            ws.cell(row=ws.max_row, column=1).font = BODY
            c = ws.cell(row=ws.max_row, column=2)
            c.font = Font(name=FONT, size=10, bold=True)
            c.number_format = "#,##0"
            c.alignment = Alignment(horizontal="right")

    # ---------------- data dictionary sheet ----------------------------------------------
    ws = wb.create_sheet("data_dictionary")
    ws.sheet_view.showGridLines = False
    ws.append(["table", "column", "definition"])
    for cell in ws[1]:
        cell.font, cell.fill = HDR_FONT, HDR_FILL
    dd_rows = ([("oligos", c, d) for c, d in OLIGO_COLUMNS]
               + [("measurements", c, d) for c, d in MEASUREMENT_COLUMNS]
               + [("modifications", c, d) for c, d in MODIFICATION_COLUMNS]
               + [("sources", c, "Provenance registry field.") for c, _ in src_cols])
    for t, c, d in dd_rows:
        ws.append([t, c, d])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = Border(bottom=THIN)
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 96
    ws.freeze_panes = "A2"

    # ---------------- data sheets ---------------------------------------------------------
    data_sheet(wb, "oligos", OLIGO_COLUMNS, oligos,
               numeric_cols=("length_nt", "n_lna", "n_moe", "n_dna", "n_5methyl_C",
                             "gap_length_nt", "flank5_len_nt", "flank3_len_nt",
                             "ps_linkage_count", "n_A", "n_C", "n_G", "n_T",
                             "gc_content_pct", "longest_g_run", "g_free_3prime_len"))
    data_sheet(wb, "measurements", MEASUREMENT_COLUMNS, meas,
               numeric_cols=("cns_tox_grade", "formulation_ca_mM", "formulation_mg_mM"))
    data_sheet(wb, "modifications", MODIFICATION_COLUMNS, mods,
               numeric_cols=("position_5to3",))
    data_sheet(wb, "sources", src_cols, srcs)

    wb.save(OUT_XLSX)
    print(f"wrote {OUT_XLSX.relative_to(ROOT)}  ({OUT_XLSX.stat().st_size / 1e6:.1f} MB)")

    # ---------------- markdown data dictionary -------------------------------------------
    lines = [f"# Data dictionary — OligoTox-CNS {VERSION}", "",
             f"Generated by `src/make_release.py` from the canonical column definitions in "
             f"`src/assemble.py`. Do not edit by hand — regenerate.", "",
             "## Missing-value tokens", "",
             "| token | meaning |", "|---|---|",
             "| `NOT_REPORTED` | the source does not report this value. Never estimated or imputed. |",
             "| `NOT_APPLICABLE` | the field has no meaning for this row. |",
             "| *(empty)* | the field does not apply to this table's row type. |", ""]
    for title, cols, path, n in (("oligos", OLIGO_COLUMNS, "data/oligos.csv", len(oligos)),
                                 ("measurements", MEASUREMENT_COLUMNS, "data/measurements.csv", len(meas)),
                                 ("modifications", MODIFICATION_COLUMNS, "data/modifications.csv", len(mods)),
                                 ("sources", src_cols, "data/sources.csv", len(srcs))):
        lines += [f"## `{title}` — {path}", "",
                  f"{n:,} rows × {len(cols)} columns.", "",
                  "| column | definition |", "|---|---|"]
        for c, d in cols:
            lines.append(f"| `{c}` | {d or 'Provenance registry field.'} |")
        lines.append("")
    OUT_MD.write_text("\n".join(lines))
    print(f"wrote {OUT_MD.relative_to(ROOT)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
