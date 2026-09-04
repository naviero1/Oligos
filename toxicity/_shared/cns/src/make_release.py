#!/usr/bin/env python3
"""Build the two release artefacts that are generated rather than written:

    deliverables/OligoTox-CNS_Dataset.xlsx   the dataset workbook
    docs/DATA_DICTIONARY.md                  the data dictionary

Both are produced from the canonical column definitions in src/assemble.py and from the CSVs
in data/, so neither can drift away from the released data.

The Summary sheet uses live formulas (COUNTA / COUNTIF) against the data sheets rather than
values computed here, so the workbook recomputes its own headline numbers when opened.
"""
from __future__ import annotations

import csv
import datetime
import pathlib
import sys

sys.path.insert(0, str(pathlib.Path(__file__).resolve().parent))
import endpoints

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

sys.path.insert(0, str(pathlib.Path(__file__).resolve().parent))
from assemble import (MEASUREMENT_COLUMNS, MODIFICATION_COLUMNS,  # noqa: E402
                      OLIGO_COLUMNS, SOURCES)

ROOT = pathlib.Path(__file__).resolve().parent.parent
DATA = ROOT / "data"
OUT_XLSX = ROOT / "deliverables" / "OligoTox-CNS_Dataset.xlsx"
OUT_MD = ROOT / "docs" / "DATA_DICTIONARY.md"

FONT = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1F3B5C")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
BODY = Font(name=FONT, size=10)
TITLE = Font(name=FONT, bold=True, size=14, color="1F3B5C")
H2 = Font(name=FONT, bold=True, size=11, color="1F3B5C")
THIN = Side(style="thin", color="D9D9D9")

TITLE_TEXT = "OligoTox-CNS — CNS toxicity of oligonucleotide therapeutics"
VERSION = "v1.0"
LICENCE = "Creative Commons Attribution 4.0 International (CC BY 4.0)"

README_LINES = [
    (TITLE_TEXT, "title"),
    (f"Dataset release {VERSION}  ·  licence: {LICENCE}", "sub"),
    ("", ""),
    ("What this is", "h2"),
    ("An open, sequence-resolved dataset of central-nervous-system toxicity measurements for "
     "oligonucleotide therapeutics, assembled for the NIH/NCATS Oligonucleotide Toxicity Open "
     "Data Challenge, Phase 2 (Data Generation).", ""),
    ("Every row pairs an oligonucleotide's design — its sequence and the position of every "
     "chemical modification in it — with a measured CNS toxicity outcome, so that a model can "
     "be trained to predict the second from the first.", ""),
    ("", ""),
    ("Sheets", "h2"),
    ("Summary            headline counts, computed live from the data sheets", "mono"),
    ("data_dictionary    every column in every table, with its definition", "mono"),
    ("oligos             one row per oligonucleotide — the predictor variables", "mono"),
    ("measurements       one row per measured outcome — the response variables", "mono"),
    ("modifications      one row per NUCLEOTIDE POSITION — the per-position chemistry", "mono"),
    ("sources            the provenance registry: citation, licence, retrieval route", "mono"),
    ("", ""),
    ("How the tables join", "h2"),
    ("measurements.oligo_id   -> oligos.oligo_id      (many measurements per oligonucleotide)", "mono"),
    ("modifications.oligo_id  -> oligos.oligo_id      (one row per position, 5' to 3')", "mono"),
    ("oligos.source_id        -> sources.source_id", "mono"),
    ("", ""),
    ("Missing values — read this before analysing", "h2"),
    ("NOT_REPORTED    the source does not report this value. It has NOT been estimated, "
     "imputed or filled from background knowledge.", ""),
    ("NOT_APPLICABLE  the field has no meaning for this row (e.g. a gap length for a vehicle "
     "control).", ""),
    ("An empty cell   the field does not apply to this table's row type.", ""),
    ("", ""),
    ("In particular: purity_pct is NOT_REPORTED for every oligonucleotide in this release. "
     "Per-compound purity values are almost never published alongside toxicity results. What "
     "IS captured, where the source states it, is the purification and identity-confirmation "
     "METHOD (purity_method, identity_confirmation). See the methodology document.", ""),
    ("", ""),
    ("A note on the Summary sheet", "h2"),
    ("The Summary sheet holds live formulas (COUNTA / COUNTIF) over the data sheets rather than "
     "values written in, so it recomputes from the data whenever the workbook is opened. Excel "
     "and LibreOffice calculate on open; a tool that reads cached values WITHOUT calculating "
     "(e.g. pandas.read_excel) will see those cells as empty. That is expected - read the CSVs "
     "in data/ for programmatic access. Each formula was independently checked against the CSVs "
     "and every one evaluates to the value qc/validate_dataset.py reports.", ""),
    ("", ""),
    ("Grades are provisional", "h2"),
    ("cns_tox_grade is an ordinal 0-3 severity assigned by the rule recorded in that row's "
     "grade_basis. Grades derived from the mouse acute tolerability score use the cut-offs "
     "published by Hagedorn et al. (2022), not thresholds invented here. All grades ship with "
     "grade_status = provisional pending subject-matter-expert review.", ""),
    ("", ""),
    ("How to cite", "h2"),
    ("Cite this dataset AND the underlying primary sources listed on the 'sources' sheet. "
     "Source H1 is CC BY 4.0; sources K1, L1 and O1 are CC BY-NC; source C1 is a US Government "
     "work in the public domain. The redistribution column on each measurement row records "
     "which terms apply to that row.", ""),
]


def read(name: str) -> list[dict]:
    """Union of the three endpoint folders (src/endpoints.py owns the split)."""
    return endpoints.load_all(name)


def data_sheet(wb, name: str, columns, rows: list[dict], numeric_cols=()):
    ws = wb.create_sheet(name)
    heads = [c for c, _ in columns]
    ws.append(heads)
    for cell in ws[1]:
        cell.font, cell.fill = HDR_FONT, HDR_FILL
        cell.alignment = Alignment(vertical="center", horizontal="left")
    for r in rows:
        ws.append([r.get(c, "") for c in heads])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY
    # numeric coercion where the column really is numeric
    for col in numeric_cols:
        if col not in heads:
            continue
        idx = heads.index(col) + 1
        for row in ws.iter_rows(min_row=2, min_col=idx, max_col=idx):
            v = row[0].value
            if isinstance(v, str) and v.replace(".", "", 1).replace("-", "", 1).isdigit():
                row[0].value = float(v) if "." in v else int(v)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(heads))}{ws.max_row}"
    for i, h in enumerate(heads, start=1):
        sample = [str(r.get(h, "")) for r in rows[:400]]
        width = max([len(h)] + [len(s) for s in sample]) + 2
        ws.column_dimensions[get_column_letter(i)].width = min(max(width, 10), 46)
    ws.sheet_view.showGridLines = False
    return ws


FIXED_ZIP_TIME = (2026, 8, 26, 0, 0, 0)


def normalise_zip_timestamps(path: pathlib.Path) -> None:
    """Rewrite an .xlsx so every archive entry carries a fixed timestamp.

    An .xlsx is a ZIP, and each entry's local header records the moment it was written, so two
    builds over identical data differ in bytes even when every cell is the same. Two things have
    to be normalised, and pinning `wb.properties` covers neither:

      1. every entry's `date_time` in the archive;
      2. `docProps/core.xml`, whose `dcterms:modified` openpyxl rewrites to the save time on the
         way out, overriding whatever the workbook properties said.

    Entry order, contents and compression are otherwise preserved exactly.
    """
    import re
    import shutil
    import tempfile
    import zipfile

    stamp = "2026-08-26T00:00:00Z"
    with zipfile.ZipFile(path) as src:
        items = []
        for i in src.infolist():
            data = src.read(i.filename)
            if i.filename == "docProps/core.xml":
                text = data.decode("utf-8")
                for field in ("created", "modified"):
                    text = re.sub(rf"(<dcterms:{field}[^>]*>)[^<]*(</dcterms:{field}>)",
                                  rf"\g<1>{stamp}\g<2>", text)
                data = text.encode("utf-8")
            items.append((i, data))
    tmp = pathlib.Path(tempfile.mkstemp(suffix=".xlsx")[1])
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as dst:
        for info, data in items:
            new_info = zipfile.ZipInfo(info.filename, date_time=FIXED_ZIP_TIME)
            new_info.compress_type = info.compress_type
            new_info.external_attr = info.external_attr
            new_info.internal_attr = info.internal_attr
            new_info.create_system = 3
            dst.writestr(new_info, data)
    shutil.move(str(tmp), str(path))


def main() -> int:
    OUT_XLSX.parent.mkdir(exist_ok=True)
    OUT_MD.parent.mkdir(exist_ok=True)
    oligos, meas, mods, srcs = (read("oligos"), read("measurements"),
                                read("modifications"), read("sources"))
    src_cols = [(k, "") for k in srcs[0]]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    # Byte-deterministic output: openpyxl stamps docProps/core.xml with the current time, so
    # two builds over identical data differ. Pin the document properties to the release date so
    # the workbook, like the PDFs, rebuilds to the same bytes.
    wb.properties.created = wb.properties.modified = datetime.datetime(2026, 8, 26, 0, 0, 0)
    wb.properties.creator = wb.properties.lastModifiedBy = "OligoTox-CNS build pipeline"
    wb.properties.title = TITLE_TEXT

    # ---------------- README -------------------------------------------------------------
    ws = wb.create_sheet("README")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 118
    for text, kind in README_LINES:
        ws.append([text])
        c = ws.cell(row=ws.max_row, column=1)
        c.font = {"title": TITLE, "h2": H2,
                  "sub": Font(name=FONT, size=10, italic=True, color="52514E"),
                  "mono": Font(name="Consolas", size=9.5)}.get(kind, BODY)
        c.alignment = Alignment(wrap_text=True, vertical="top")

    # ---------------- Summary (live formulas) --------------------------------------------
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 56
    ws.column_dimensions["B"].width = 16
    ws.append([TITLE_TEXT]); ws["A1"].font = TITLE
    ws.append([f"Release {VERSION}. Every figure below is a formula over the data sheets."])
    ws["A2"].font = Font(name=FONT, size=10, italic=True, color="52514E")
    ws.append([])

    g_col = get_column_letter([c for c, _ in MEASUREMENT_COLUMNS].index("cns_tox_grade") + 1)
    h_col = get_column_letter([c for c, _ in MEASUREMENT_COLUMNS].index("is_human_system") + 1)
    s_col = get_column_letter([c for c, _ in MEASUREMENT_COLUMNS].index("study_type") + 1)
    b_col = get_column_letter([c for c, _ in OLIGO_COLUMNS].index("modification_position_basis") + 1)
    p_col = get_column_letter([c for c, _ in OLIGO_COLUMNS].index("purity_pct") + 1)

    # Bounded ranges, not whole-column references: whole-column COUNTIF over a workbook
    # containing a 32,569-row sheet takes LibreOffice more than nine minutes to recalculate.
    NO = len(oligos) + 200          # oligos data rows + headroom
    NM = len(meas) + 200            # measurements data rows + headroom
    ND = len(mods) + 200            # modifications data rows + headroom
    NS = len(srcs) + 20
    orng = lambda c: f"oligos!${c}$2:${c}${NO}"
    mrng = lambda c: f"measurements!${c}$2:${c}${NM}"
    seq_col = get_column_letter([c for c, _ in OLIGO_COLUMNS].index("sequence_base") + 1)

    rows = [
        ("Counts", None),
        ("Oligonucleotides", f'=COUNTA({orng("A")})'),
        ("CNS toxicity measurements", f'=COUNTA({mrng("A")})'),
        ("Per-position modification records", f'=COUNTA(modifications!$A$2:$A${ND})'),
        ("Distinct sources", f'=COUNTA(sources!$A$2:$A${NS})'),
        (None, None),
        ("Coverage", None),
        ("Oligonucleotides with a published sequence",
         f'=COUNTA({orng("A")})-COUNTIF({orng(seq_col)},"NOT_REPORTED")'),
        ("Oligonucleotides with position-resolved modifications",
         f'=COUNTIF({orng(b_col)},"position_resolved*")'),
        ("Oligonucleotides with a reported purity value",
         f'=COUNTA({orng("A")})-COUNTIF({orng(p_col)},"NOT_REPORTED")'),
        (None, None),
        ("Measurements by study type", None),
        ("in vitro", f'=COUNTIF({mrng(s_col)},"in_vitro")'),
        ("in vivo (animal)", f'=COUNTIF({mrng(s_col)},"animal_invivo")'),
        ("clinical (human)", f'=COUNTIF({mrng(s_col)},"clinical")'),
        ("measured in a human or human-derived system",
         f'=COUNTIF({mrng(h_col)},"TRUE")'),
        (None, None),
        ("Severity grade distribution", None),
        ("grade 0 — no observable CNS signal", f'=COUNTIF({mrng(g_col)},0)'),
        ("grade 1 — mild", f'=COUNTIF({mrng(g_col)},1)'),
        ("grade 2 — moderate", f'=COUNTIF({mrng(g_col)},2)'),
        ("grade 3 — severe", f'=COUNTIF({mrng(g_col)},3)'),
    ]
    for label, formula in rows:
        if label is None:
            ws.append([])
            continue
        if formula is None:
            ws.append([label])
            ws.cell(row=ws.max_row, column=1).font = H2
        else:
            ws.append([label, formula])
            ws.cell(row=ws.max_row, column=1).font = BODY
            c = ws.cell(row=ws.max_row, column=2)
            c.font = Font(name=FONT, size=10, bold=True)
            c.number_format = "#,##0"
            c.alignment = Alignment(horizontal="right")

    # ---------------- data dictionary sheet ----------------------------------------------
    ws = wb.create_sheet("data_dictionary")
    ws.sheet_view.showGridLines = False
    ws.append(["table", "column", "definition"])
    for cell in ws[1]:
        cell.font, cell.fill = HDR_FONT, HDR_FILL
    dd_rows = ([("oligos", c, d) for c, d in OLIGO_COLUMNS]
               + [("measurements", c, d) for c, d in MEASUREMENT_COLUMNS]
               + [("modifications", c, d) for c, d in MODIFICATION_COLUMNS]
               + [("sources", c, "Provenance registry field.") for c, _ in src_cols])
    for t, c, d in dd_rows:
        ws.append([t, c, d])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = Border(bottom=THIN)
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 96
    ws.freeze_panes = "A2"

    # ---------------- human_measurements ---------------------------------------------------
    # The human rows are the most important subset, so they get their own sheet with the
    # oligonucleotide's SEQUENCE and CHEMISTRY joined inline and the TOXICITY GRADE placed up
    # front - rather than requiring a join against oligos.csv to see either.
    by_oligo = {o["oligo_id"]: o for o in oligos}
    HUMAN_FRONT = [
        ("oligo_name", "Oligonucleotide"),
        ("sequence_5to3_asprinted", "Sequence (5'->3', as printed)"),
        ("sequence_base", "Sequence (bases only)"),
        ("length_nt", "Length (nt)"),
        ("backbone_chemistry", "Backbone"),
        ("sugar_modifications", "Sugar modifications"),
        ("modification_pattern", "Modification pattern"),
        ("modification_positions", "Modification positions (5'->3')"),
        ("modification_position_basis", "Position basis"),
    ]
    hm_cols = ([("cns_tox_grade", "Toxicity grade (0-3)"), ("grade_basis", "How the grade was assigned")]
               + HUMAN_FRONT
               + [(c, c) for c, _ in MEASUREMENT_COLUMNS if c != "cns_tox_grade" and c != "grade_basis"])
    human_rows = []
    for m in meas:
        if m.get("subject_group") != "human":
            continue
        o = by_oligo.get(m["oligo_id"], {})
        row = {}
        for key, _ in hm_cols:
            row[key] = o.get(key, "") if key in dict(HUMAN_FRONT) else m.get(key, "")
        human_rows.append(row)
    ws = data_sheet(wb, "human_measurements", [(k, "") for k, _ in hm_cols], human_rows,
                    numeric_cols=("cns_tox_grade", "length_nt"))
    for i, (_, label) in enumerate(hm_cols, start=1):
        ws.cell(row=1, column=i).value = label
    with_seq = sum(1 for r in human_rows if r["sequence_5to3_asprinted"] not in ("NOT_REPORTED", ""))
    print(f"  human_measurements: {len(human_rows)} rows, {with_seq} carry a sequence "
          f"({100 * with_seq / max(len(human_rows), 1):.0f}%)")

    # ---------------- German's analysis ----------------------------------------------------
    # One row per oligonucleotide: what it is, its sequence, what was done to that sequence, and
    # how toxic it turned out. Deliberately narrow - the modelling view, not the provenance view.
    GRADE_LABEL = {"0": "0 - none", "1": "1 - mild", "2": "2 - moderate", "3": "3 - severe"}
    per_oligo = {}
    for m in meas:
        g = m.get("cns_tox_grade", "")
        d = per_oligo.setdefault(m["oligo_id"], {"n": 0, "worst": None, "classes": set(),
                                                 "readouts": set()})
        d["n"] += 1
        d["classes"].add(m.get("subject_class", ""))
        d["readouts"].add(m.get("readout_name", ""))
        if g != "" and (d["worst"] is None or int(g) > d["worst"]):
            d["worst"] = int(g)
    ga_rows = []
    for o in oligos:
        d = per_oligo.get(o["oligo_id"], {"n": 0, "worst": None, "classes": set(), "readouts": set()})
        # NB: named mod_text, not mods -- `mods` is the modifications table in the enclosing
        # scope and shadowing it silently corrupted that sheet.
        mod_text = o.get("modification_pattern", "")
        pos = o.get("modification_positions", "")
        if pos not in ("", "NOT_REPORTED"):
            mod_text = (f"{mod_text} | per-position: {pos}"
                        if mod_text not in ("", "NOT_REPORTED") else f"per-position: {pos}")
        ga_rows.append({
            "oligo_id": o["oligo_id"],
            "oligo": o["oligo_name"],
            "sequence": o.get("sequence_5to3_asprinted", "NOT_REPORTED"),
            "modification": mod_text or "NOT_REPORTED",
            "toxicity": GRADE_LABEL.get(str(d["worst"]), "not graded") if d["worst"] is not None else "not graded",
            "toxicity_grade_numeric": d["worst"] if d["worst"] is not None else "",
            "n_measurements": d["n"],
            "subject_class": "; ".join(sorted(c for c in d["classes"] if c)),
            "readouts": "; ".join(sorted(r for r in d["readouts"] if r))[:250],
            "source_id": o.get("source_id", ""),
        })
    ga_rows.sort(key=lambda r: (r["sequence"] in ("NOT_REPORTED", ""),
                                -(r["toxicity_grade_numeric"] if isinstance(r["toxicity_grade_numeric"], int) else -1),
                                r["oligo_id"]))
    ga_cols = [("oligo_id", ""), ("oligo", ""), ("sequence", ""), ("modification", ""),
               ("toxicity", ""), ("toxicity_grade_numeric", ""), ("n_measurements", ""),
               ("subject_class", ""), ("readouts", ""), ("source_id", "")]
    ws = data_sheet(wb, "German's analysis", ga_cols, ga_rows,
                    numeric_cols=("toxicity_grade_numeric", "n_measurements"))
    LABELS = ["Oligo ID", "Oligonucleotide", "Sequence (5'->3')", "Modification to that sequence",
              "Toxicity", "Toxicity grade", "Measurements", "Subject class", "Readouts", "Source"]
    for i, lab in enumerate(LABELS, start=1):
        ws.cell(row=1, column=i).value = lab
    gs = sum(1 for r in ga_rows if r["sequence"] not in ("NOT_REPORTED", ""))
    print(f"  German's analysis: {len(ga_rows)} oligonucleotides, {gs} with a sequence, "
          f"{sum(1 for r in ga_rows if r['toxicity'] != 'not graded')} graded")

    # ---------------- data sheets ---------------------------------------------------------
    data_sheet(wb, "oligos", OLIGO_COLUMNS, oligos,
               numeric_cols=("length_nt", "n_lna", "n_moe", "n_dna", "n_5methyl_C",
                             "gap_length_nt", "flank5_len_nt", "flank3_len_nt",
                             "ps_linkage_count", "n_A", "n_C", "n_G", "n_T",
                             "gc_content_pct", "longest_g_run", "g_free_3prime_len"))
    data_sheet(wb, "measurements", MEASUREMENT_COLUMNS, meas,
               numeric_cols=("cns_tox_grade", "formulation_ca_mM", "formulation_mg_mM"))
    data_sheet(wb, "modifications", MODIFICATION_COLUMNS, mods,
               numeric_cols=("position_5to3",))
    data_sheet(wb, "sources", src_cols, srcs)

    wb.save(OUT_XLSX)
    normalise_zip_timestamps(OUT_XLSX)
    print(f"wrote {OUT_XLSX.relative_to(ROOT)}  ({OUT_XLSX.stat().st_size / 1e6:.1f} MB)")

    # ---------------- markdown data dictionary -------------------------------------------
    lines = [f"# Data dictionary — OligoTox-CNS {VERSION}", "",
             f"Generated by `src/make_release.py` from the canonical column definitions in "
             f"`src/assemble.py`. Do not edit by hand — regenerate.", "",
             "## Missing-value tokens", "",
             "| token | meaning |", "|---|---|",
             "| `NOT_REPORTED` | the source does not report this value. Never estimated or imputed. |",
             "| `NOT_APPLICABLE` | the field has no meaning for this row. |",
             "| *(empty)* | the field does not apply to this table's row type. |", ""]
    for title, cols, path, n in (("oligos", OLIGO_COLUMNS, "data/oligos.csv", len(oligos)),
                                 ("measurements", MEASUREMENT_COLUMNS, "data/measurements.csv", len(meas)),
                                 ("modifications", MODIFICATION_COLUMNS, "data/modifications.csv", len(mods)),
                                 ("sources", src_cols, "data/sources.csv", len(srcs))):
        lines += [f"## `{title}` — {path}", "",
                  f"{n:,} rows × {len(cols)} columns.", "",
                  "| column | definition |", "|---|---|"]
        for c, d in cols:
            # Escape pipes: many definitions are enumerations written "a | b | c", and an
            # unescaped pipe splits the markdown row into extra cells. 18 rows rendered wrong
            # before this was caught.
            text = (d or "Provenance registry field.").replace("|", "\\|")
            row = f"| `{c}` | {text} |"
            assert row.count("|") - row.count("\\|") == 3, f"malformed dictionary row for {c}"
            lines.append(row)
        lines.append("")
    OUT_MD.write_text("\n".join(lines))
    print(f"wrote {OUT_MD.relative_to(ROOT)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
