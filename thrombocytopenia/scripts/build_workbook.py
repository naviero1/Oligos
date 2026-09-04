#!/usr/bin/env python3
"""Build the single-file reviewer workbook for the thrombocytopenia dataset.

Phase 2 allows the dataset to be supplied "either by including a data file in Excel
(or similar format) or a document with instructions ... on how to access and download
the raw data". This produces the former: every table on its own sheet, filtered,
frozen and column-sized, fronted by a README sheet. The canonical CSVs in the
repository remain the source of truth; this workbook is generated from them.

Sheet order is deliberate — the sheets a reviewer needs first come first:
  README                 orientation, rubric, and the two modelling cautions
  Germans_analysis       compound · sequence · modification · toxicity, worst first
  measurements_human     THE priority subset, denormalised so sequence and grade
                         sit beside every row without a join
  bridge_human_animal    the 23 compounds characterised in BOTH systems
  measurements_animal    the animal side, same shape as human
  oligos                 canonical compound table
  measurements           canonical measurement table
  merged_analysis_view   full denormalised join

Usage:  python3 scripts/build_workbook.py
"""
import csv, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ENDPOINT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BASE = os.path.join(ENDPOINT, "data")
OUT = os.path.join(ENDPOINT, "submission", "OligoTox-Thrombocytopenia_dataset.xlsx")

HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(color="FFFFFF", bold=True, size=9)
# grade colouring: a reviewer should see severity without reading numbers
GRADE_FILL = {"0": PatternFill("solid", fgColor="E8F3E8"),
              "1": PatternFill("solid", fgColor="FFF6E0"),
              "2": PatternFill("solid", fgColor="FDE7D6"),
              "3": PatternFill("solid", fgColor="F8D4D4")}

SHEETS = [
    ("Germans_analysis", "germans_analysis.csv",
     "One row per compound: what the molecule IS (sequence and modifications) and what it DID "
     "(toxicity). Sorted worst-first. Human evidence is broken out separately because averaging "
     "it into the animal data would hide the subset that matters most."),
    ("measurements_human", "measurements_human.csv",
     "THE PRIORITY SUBSET — every human row, denormalised so the sequence and the toxicity grade "
     "sit beside each measurement without needing a join. Columns 4 and 5 are sequence and grade."),
    ("bridge_human_animal", "bridge_human_animal.csv",
     "Compounds characterised in BOTH human and animal systems — the set a cross-species model can "
     "actually be trained and validated on. Holding human rows and animal rows separately does not "
     "demonstrate extrapolation; this does."),
    ("measurements_animal", "measurements_animal.csv",
     "Animal rows, same denormalised shape as the human sheet."),
    ("oligos", "oligos.csv", None),
    ("measurements", "measurements.csv", None),
    ("merged_analysis_view", "oligotox_thrombo_merged.csv", None),
]


def add_sheet(wb, title, path, note):
    with open(os.path.join(BASE, path), newline="", encoding="utf-8") as f:
        rows = list(csv.reader(f))
    if not rows:
        return 0
    ws = wb.create_sheet(title[:31])
    r0 = 1
    if note:
        c = ws.cell(1, 1, note)
        c.font = Font(italic=True, size=9, color="555555")
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(10, len(rows[0])))
        ws.row_dimensions[1].height = 30
        r0 = 3

    header = rows[0]
    gi = header.index("thrombocytopenia_grade") + 1 if "thrombocytopenia_grade" in header else None
    mi = header.index("max_toxicity_grade") + 1 if "max_toxicity_grade" in header else None

    for i, row in enumerate(rows):
        for j, v in enumerate(row):
            cell = ws.cell(r0 + i, j + 1, v)
            if i == 0:
                cell.fill = HDR_FILL
                cell.font = HDR_FONT
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            elif (gi and j + 1 == gi) or (mi and j + 1 == mi):
                cell.fill = GRADE_FILL.get(v.strip(), PatternFill())
                cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = ws.cell(r0 + 1, 4 if title.startswith(("measurements_h", "measurements_a")) else 2)
    for j, h in enumerate(header, 1):
        widths = [len(str(r[j - 1])) for r in rows[1:500] if j - 1 < len(r)]
        ws.column_dimensions[get_column_letter(j)].width = \
            max(len(h) + 2, min(42, (max(widths) if widths else 10) + 2))
    ws.auto_filter.ref = f"A{r0}:{get_column_letter(len(header))}{r0 + len(rows) - 1}"
    return len(rows) - 1


def main():
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("README")
    counts = {}
    info = [
        ("OligoTox-Thrombocytopenia — Phase 2 dataset", ""),
        ("", ""),
        ("Endpoint", "thrombocytopenia / platelet toxicity (this workbook contains no other endpoint's data)"),
        ("Licence", "CC-BY 4.0 — rights are ALSO tracked per row in the redistribution column"),
        ("Repository", "naviero1/Oligos → thrombocytopenia/"),
        ("", ""),
        ("Sheet", "Contents"),
        ("Germans_analysis", "One row per compound: oligo · sequence · modification · toxicity. Sorted worst-first, with human evidence broken out."),
        ("measurements_human", "THE PRIORITY SUBSET. Every human row with the sequence and toxicity grade joined in (columns 4 and 5) — no join required."),
        ("bridge_human_animal", "Compounds characterised in BOTH human and animal systems — the cross-species extrapolation set."),
        ("measurements_animal", "Animal rows, same denormalised shape."),
        ("oligos", "Canonical compound table — identity and design predictors."),
        ("measurements", "Canonical measurement table — outcomes and context, normalised."),
        ("merged_analysis_view", "Full denormalised join of both canonical tables."),
        ("", ""),
        ("Grade rubric", "0 = no signal at tested exposure · 1 = mild / reversible (≥100×10⁹/L, or activation only above clinical concentration) · 2 = moderate (50–99×10⁹/L, or activation at clinically relevant concentration) · 3 = severe (<50×10⁹/L, CTCAE grade 4, antibody-mediated, bleeding attributable to thrombocytopenia, or discontinuation)"),
        ("Two grading conventions", "An INCIDENCE row is graded by the SEVERITY OF THE EVENT, not by how small the incidence is. BLEEDING is graded on attribution — grade 3 requires haemorrhage attributable to thrombocytopenia."),
        ("Missing values", "the literal string TBD — never guessed, never imputed as zero. NA means the field does not apply."),
        ("", ""),
        ("⚠ Before modelling — 1", "EXCLUDE dose_or_conc_value == '0'. Those are control/placebo arms; some carry grade 1–2 correctly (a placebo subject really did drop below threshold), but joined naively they teach a model that a compound acts at zero dose."),
        ("⚠ Before modelling — 2", "ACCOUNT FOR study_type. Severe events are observed in trials, not in dishes, so grade is partly confounded with study design. Do not let a model learn that as biology."),
        ("⚠ Before modelling — 3", "imetelstat's thrombocytopenia is on-target telomerase myelosuppression, NOT phosphorothioate platelet binding. Exclude it from structure-activity analysis."),
        ("", ""),
        ("Verification", "Rows verified against their primary source carry 'verified_against_source' in notes. The dataset is PARTIALLY verified — see METHODOLOGY.md for which blocks are complete."),
        ("Purity data", "Largely unavailable and never inferred: this is a curation of published results for third-party compounds, and sources rarely report purity for the material they tested. See the methodology document."),
        ("Full documentation", "schema.md · METHODOLOGY.md · SOURCES.md · STATUS.md · submission/narrative.pdf"),
    ]
    for i, (a, b) in enumerate(info, 1):
        ca = ws.cell(i, 1, a)
        ca.font = Font(bold=(i == 1 or a in ("Sheet",) or a.startswith("⚠")),
                       size=(13 if i == 1 else 9),
                       color=("A33000" if a.startswith("⚠") else "000000"))
        ca.alignment = Alignment(vertical="top")
        cb = ws.cell(i, 2, b)
        cb.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 112

    for title, path, note in SHEETS:
        counts[title] = add_sheet(wb, title, path, note)

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    print(f"wrote {os.path.relpath(OUT, ENDPOINT)}  ({os.path.getsize(OUT)/1024/1024:.1f} MB)")
    for k, v in counts.items():
        print(f"  {k:<24} {v:>5} rows")


if __name__ == "__main__":
    main()
